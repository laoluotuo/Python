#coding:utf-8
"""
全国企业信用公示系统江苏站的自动化查询工具. 是本人接了一个兼职查数据的活, 用来练手的.
功能: 读取Excel表格中的公司名, 做查询, 屏幕显示结果并插入原表格. 
原理: 在IPython Notebook中运行, 自动读取表格中的公司名, 访问http://www.jsgsj.gov.cn:58888/province/网站, 人工输入图形验证码, 自动给出符合条件的
公司数据, 并插入原Excel表. 同时显示. 可指定跳到的Excel行查询, 可整理公司名字段, 自动修正. 
范例:
---------------------------------------------------------------
公司名:江苏XX医疗器械有限公司
共有1项查询结果
******************************************************************************
链接地址:
http://www.jsgsj.gov.cn:58888/ecipplatform/publicInfoQueryServlet.json?pageView=true&org=8131E47F95953FD6DE84A218637DDB2C&id=D02D081CBE2B27081760B79B0BFDE9FF&seqId=55CD45F0906EC0947371AAB2AA1DA5C8&tmp=51
913212915855760989   ||  江苏XX医疗器械有限公司   ||  有限责任公司
   ||  XXX   ||  800万元人民币   ||  2011年11月16日   ||  泰州市药城大道XXX号医疗器械区一期综合楼XXX号.   ||  2011年11月16日   ||  2041年11月15日   ||  批发二类医疗器械、三类医疗器械、体外诊断试剂（器械类），一类医疗器械生产销售。（依法须经批准的项目，经相关部门批准后方可开展经营活动）   ||  泰州医药高新技术产业开发区市场监督管理局   ||  2016年10月24日   ||  在业   ||  空   ||  
----------------------------------------------------------------

"""


import requests
import json
import time
from openpyxl import load_workbook
from IPython.display import Image, display
"""
江苏站
"""
class colldata():
    def __init__(self):
        self.req = requests.session()
        self.typename = None
        self.result = None
        self.items = None
        self.shortinfo = {}
        self.longinfo = [] #字典格式详细信息
        self.finaldata = []#列表格式详细信息, 内容与上面相同
        self.nums = [1]
        self.wb = load_workbook('addr.xlsx')
        self.ws = self.wb['Sheet1']
        self.ws_temp = self.wb['Sheet2']
        self.cur = 'b3'
        self.order = ['REG_NO', 'CORP_NAME', 'ZJ_ECON_KIND', 'OPER_MAN_NAME', 'REG_CAPI', 'START_DATE', 'ADDR', 'FARE_TERM_START', 'FARE_TERM_END', 'FARE_SCOPE', 'BELONG_ORG', 'CHECK_DATE', 'CORP_STATUS', 'WRITEOFF_DATE']
        self.header = {
         'Referer': 'http://www.jsgsj.gov.cn:58888/ecipplatform/jiangsu.jsp?typeName=43CBEF84ED4F37C8851664629C93EA6C0855DAD80F01ECABCCD74CE6B8869EBA&searchType=qyxx',
         'Origin': 'http://www.jsgsj.gov.cn:58888',
         'Host': 'www.jsgsj.gov.cn:58888',
         'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/54.0.2840.71 Safari/537.36',
}
        self.data = {
            'name':'有限公司'
        }
        self.submit = {
            'name':self.typename,
            'searchType':'qyxx',
            'pageNo':1,
            'pageSize':10

        }


    def timestamp(self):
        stamp = str(time.time() * 1000)[0:13]
        return stamp


    def verify(self):
        re_scode = self.req.get('http://www.jsgsj.gov.cn:58888/ecipplatform/rand_img.jsp?type=3&temp=' + self.timestamp(),headers=self.header, stream=True, verify=False)
        display(Image(re_scode.content))
        #输入验证码并请求
        vericode = input('输入验证码中要求的结果: ')
        if vericode in ('stop', 'STOP'):
            return 'stop'
        elif vericode.startswith('jump'):
            return vericode
        self.data['verifyCode'] = vericode
        self.data['type'] = 3
        self.data['temp'] = self.timestamp()
        re_page1 = self.req.post('http://www.jsgsj.gov.cn:58888/ecipplatform/infoQueryServlet.json?checkCode=true', headers=self.header, data=self.data, stream=True, verify=False)
        try:
            self.typename = json.loads(re_page1.text.strip('[]'))['bean']['name']
#             print('验证码返回: ',re_page1.text)
        except KeyError:
            print('验证码错误? ')
            self.verify()


    def itemck(self):
        """
        #请求返回页和条目的页面, total是项目数, totalpage是页数
        """
        data_check = {
        'keyWord':self.typename,
        'searchType':'qyxx'
    }
        re_check = self.req.post('http://www.jsgsj.gov.cn:58888/ecipplatform/infoQueryServlet.json?queryCinfoCount=true', headers=self.header, data=data_check, stream=True, verify=False)
        self.items = json.loads(re_check.text.strip('[]'))
        if self.items.get('ERROR'):
            print(self.items['ERROR'])
            return 'ERROR'
        elif self.items['total'] == 0:
            print('没有查到数据 ')
            return 'NODATA'

        else:
            print('共有{}项查询结果'.format(self.items['total']))


    def getlinks(self):
        """
        拼接最终页面链接(有几个拼几个), 并连接取得最终数据(一秒发送一次请求)
        """
        self.submit['name'] = self.typename
        getjsn_page = self.req.post('http://www.jsgsj.gov.cn:58888/ecipplatform/infoQueryServlet.json?queryCinfo=true', headers=self.header, data=self.submit, stream=True, verify=False)
        self.result = getjsn_page.text
        for item in json.loads(self.result)['items']: # 拼接并屏显当前的链接详情
            time.sleep(1)
            try:
                self.shortinfo = item
            except IndexError:
                print('无此相关数据!')
            link = 'http://www.jsgsj.gov.cn:58888/ecipplatform/publicInfoQueryServlet.json?pageView=true&' + 'org=' + self.shortinfo['ORG'] + '&id=' + self.shortinfo['ID'] + '&seqId=' + self.shortinfo['SEQ_ID'] + '&tmp=51'

            getlink = self.req.get(link, headers=self.header)
            longinfo = json.loads(getlink.text.strip('[]'))
            self.longinfo.append(longinfo)

            print('******************************************************************************')
            print('链接地址:\n{}'.format(link))
            self.finaldata_output(longinfo)   #取出数据, 放入self.finaldata


    def dataask(self, ):
        """
        数据插入
        """
        condition = 1
        while condition == 1: #循环求输入要插入的数据条目, 得到self.nums这个列表,与self.finaldata对应
            if len(self.finaldata) > 1:
                items = input('目前查询结果超过一条,请给出你的选择,从1到n,用空格隔开: ')
                self.nums = items.split()
                try:
                    for i in self.nums:
                        if (not isinstance(int(i), int)) or int(i) > len(self.finaldata):
                            print('请输入整数,> 0且 <= 数据条数.')
                            break
                        else:
                            print('你选择的数据条目是: {}'.format(self.nums))
                            self.nums = list(map(int, self.nums))
                            condition = 0
                except TypeError:
                    print('请输入整数')
            elif len(self.finaldata) == 0:
                return
            else:
                condition = 0


    def finaldata_output(self, longinfo):
        """
        将返回的longinfo, 输出到屏幕
        """
        lst = []
        count = 0
        for _ in range(14):
            if longinfo[self.order[count]] is None:
                longinfo[self.order[count]] = ''
            print(longinfo[self.order[count]], '  ||  ', end='')
            lst.append(longinfo[self.order[count]])
            count += 1
        print('\n')
        self.finaldata.append(lst)


    def get_sleced(self):
        """
        把查询到的最终数据插入表. 若只有一条,插入主表; 若有多条,第一条插入主表, 其他插入临时表Sheet2
        """
        if len(self.nums) == 1 and self.nums[0] == 1:
            self.handle_cell(self.ws, self.cur)
        elif self.nums[0] == 000 :
            self.handle_cell(self.ws, self.cur)
            tempcur = self.cur
            for _ in range(self.items['total'] - 1):
                self.handle_cell(self.ws_temp, tempcur)
                tempcur = tempcur[0] + str(int(tempcur[1:]) +1)
        else:
            self.handle_cell(self.ws, self.cur)
            tempcur = self.cur
            for _ in range(len(self.nums) - 1):
                self.handle_cell(self.ws_temp, tempcur)
                tempcur = tempcur[0] + str(int(tempcur[1:]) +1)
        self.wb.save('addr.xlsx')
        self.reset() #重置所有计数器, 防止污染
        self.cur = self.cur[0] + str(int(self.cur[1:]) +1)


    def handle_cell(self,table, head):
        """
        处理某表格的某行, 插入给出链接的所有字段
        """
        row = 'e' + head.lstrip('b')
        col = 'r' + head.lstrip('b')
        data = self.finaldata.pop(0)
        if len(data) < 14:
            for _ in range(14 - len(data)):
                data.append('')
        #循环读取self.order中的字段, 用来从finaldata的字典中取出数据, 依次放入单元格.
        counter = 0
        for i in table[row:col]:
            for j in i:
                j.value = data[counter]
                counter += 1


    def wb_handle(self, jump=None):
        """
        读取EXCEL表, 取得公司名, 交给handle_str处理并赋值给self.data['name']
        """
        if jump is not None:
            self.ws[self.cur] = string = self.ws[jump].value
            self.cur = jump
        else:
            string = self.ws[self.cur].value
        self.handle_str(string)
        print('===========================================================================')
        print('原公司名:{}, 现公司名:{}'.format(string, self.data['name']))


    def handle_str(self, string):
        value = string
        spec = ['(自配送)',
                '（配送）',
                '(经营)',
                '(配送)',
                '（经营）',
                '(自配)'
                ]
        if value.startswith('HP'):
            value = value.lstrip('HP')
            # print(j)
        elif value.startswith('PS'):
            value = value.lstrip('PS')
            # print(j)
        for k in spec:
            if k in value:
                value = value.rstrip(k)
                continue
            elif value.endswith('作废'):
                print('作废条目: ', value)
                continue
        self.data['name'] = value


    def reset(self): #重置关键计数器, 防止数据出错.
        self.result = None
        self.items = None
        self.shortinfo = {}
        self.longinfo = []
        self.finaldata = []
        self.nums = [1]


if __name__ == '__main__':

    def main():
        con = colldata()
        while True:
            con.wb_handle()
            re = con.verify()
            if re == 'stop':
                print('程序中止')
                print('当前查询行: {}'.format(con.cur))
                return
            elif re is not None:
                if re.startswith('jump'):
                    jump = re.lstrip('jump')
                    con.wb_handle(jump)
                    con.verify()
            dataexist = con.itemck()
            if dataexist == 'ERROR' :
                print('当前查询列: {}'.format(con.cur))
                return
            elif dataexist == 'NODATA':
                con.cur = con.cur[0] + str(int(con.cur[1:]) +1)
                continue
            con.getlinks()
            con.dataask()
            con.get_sleced()


    main()

